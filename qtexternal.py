import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QInputDialog,
                             QLineEdit, QScrollArea, QFrame, QDialog, QMessageBox, QShortcut, QSplitter, QTextEdit)
from PyQt5.QtCore import Qt, QTimer, QSize, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QKeySequence, QPainter, QColor, QImage
import datetime
from openpyxl import Workbook, load_workbook
import os
from printer import *
import socket
import json
import select


class Order:
    order_counter = 1
    food_items_info = {
        "Friikad": {"price": 2, "key": "1"},
        "Friikartulid vinkudega": {"price": 3, "key": "2"},
        "Bataadifriikad": {"price": 3, "key": "3"},
        "Bataadifriikad vinkudega": {"price": 4, "key": "4"},
        "Pelmeenid": {"price": 2.5, "key": "5"},
        "Burger": {"price": 3, "key": "6"},

    }

    drin_items = {
        "Coca-Cola": {"price": 1.5, "key": "g"},
        "Fanta": {"price": 1.5, "key": "h"},
        "Limpa": {"price": 1.5, "key": "j"},
        "Vesi": {"price": 1.5, "key": "k"},
        "Kokteil klassika": {"price": 2, "key": "l"},
        "Kokteil saladus": {"price": 2, "key": ";"},
    }

    bar_items_info = {
        "hammertime": {"price": 4, "key": "z"},
        "akutrell": {"price": 4, "key": "x"},
        "rohtlakiisu": {"price": 4, "key": "v"},
        "kruvikeeraja": {"price": 4, "key": "b"},
        "kahemehesaag": {"price": 6, "key": "n"},
        "relakas": {"price": 2, "key": "m"},
        "Lammutaja BirgIT": {"price": 4, "key": ","},
        "Lendav sirel": {"price": 4, "key": "."},
        "Tugevusõpetus": {"price": 4, "key": "/"},
        "red bull": {"price": 2, "key": "'"}
    }

    combined_items_info = food_items_info | drin_items | bar_items_info

    def __init__(self, order_number=None, food_items=None, drink_items=None,
                 bar_items=None, customer_name=None, creation_time=None,
                 total_cost=0, status="active", card_or_cash="card"):
        if order_number:
            self.order_number = order_number
            Order.order_counter = order_number
        else:
            self.order_number = Order.order_counter
        Order.order_counter += 1

        self.creation_time = creation_time if creation_time else datetime.datetime.now()
        self.food_items = food_items if food_items else {}
        self.bar_items = bar_items if bar_items else {}
        self.drink_items = drink_items if drink_items else {}
        self.combined_order_items = self.food_items | self.bar_items
        self.customer_name = customer_name
        self.total_cost = total_cost
        self.status = status
        self.card_or_cash = card_or_cash

    @classmethod
    def get_item_price(cls, item_name):
        if item_name in cls.combined_items_info:
            return cls.combined_items_info.get(item_name, {}).get("price", 0)

    @classmethod
    def get_item_key(cls, item_name):
        if item_name in cls.combined_items_info:
            return cls.combined_items_info.get(item_name, {}).get("key", "")

    @classmethod
    def get_item_by_key(cls, key):
        for item, details in cls.combined_items_info.items():
            if details["key"] == key:
                return item
        return None

    @classmethod
    def calculate_total_price(cls, order):
        return sum(cls.get_item_price(item) * quantity for item, quantity in order.combined_order_items.items())

    @classmethod
    def get_items_display(cls, order):
        return "Current Order: " + ", ".join(f"{item} x{count}" for item, count in order.combined_order_items.items())

    @classmethod
    def add_item(cls, item, order):
        if item in Order.food_items_info:
            if item in order.food_items:
                order.food_items[item] += 1
            else:
                order.food_items[item] = 1

        elif item in Order.bar_items_info:
            if item in order.bar_items:
                order.bar_items[item] += 1
            else:
                order.bar_items[item] = 1

        order.combined_order_items = order.food_items | order.bar_items
        return order

    @classmethod
    def remove_item(cls, item, order):
        if item in Order.food_items_info:
            if item in order.food_items:
                if order.food_items[item] > 1:
                    order.food_items[item] -= 1
                else:
                    del order.food_items[item]

        elif item in Order.bar_items_info:
            if item in order.bar_items:
                if order.bar_items[item] > 1:
                    order.bar_items[item] -= 1
                else:
                    del order.bar_items[item]

        order.combined_order_items = order.food_items | order.bar_items

    @classmethod
    def get_prompt_text(cls):
        food_text = ", ".join(
            f"'{details['key']}' {item}" for item, details in Order.food_items_info.items())
        bar_text = ", ".join(
            f"'{details['key']}' {item}" for item, details in Order.bar_items_info.items())
        return food_text + '\n' + bar_text + '\nEnter: finish order'


class KitchenManagerApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Kitchen Manager")
        self.setGeometry(400, 400, 800, 800)

        self.orders_list = []

        self.current_order = None
        self.new_order = None

        self.rpi_ip = '100.118.0.128'
        self.rpi_port = 5000
        self.rpi_connected = False

        # Create instances of KitchenWindow and CustomerWindow
        self.kitchen_window = KitchenWindow(self)
        self.kitchen_window.show()

        self.load_or_create_workbook()

        self.init_ui()

        # Start periodic connection check
        self.connection_timer = QTimer(self)
        self.connection_timer.timeout.connect(self.start_ping)
        self.connection_timer.start(5000)  # Check every 5 seconds

    def init_ui(self):
        splitter = QSplitter(Qt.Horizontal)
        self.setCentralWidget(splitter)

        self.main_content_widget = QWidget()
        self.main_layout = QVBoxLayout(self.main_content_widget)

        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        self.info_widget = QTextEdit()
        self.info_widget.setReadOnly(True)
        self.info_widget.setText("Information Panel")

        self.shortcuts_widget = QTextEdit()
        self.shortcuts_widget.setReadOnly(True)
        self.shortcuts_widget.setText("Key Shortcuts")

        right_layout.addWidget(self.info_widget)
        right_layout.addWidget(self.shortcuts_widget)

        splitter.addWidget(self.main_content_widget)
        splitter.addWidget(right_panel)
        splitter.setSizes([60, 40])

        self.set_shortcuts()
        self.update_info_widget()
        self.update_shortcuts_widget()
        self.update_displays()

    def set_shortcuts(self):
        # Set up key bindings
        QShortcut(QKeySequence('N'), self, self.new_order_process)
        QShortcut(QKeySequence('U'), self, self.update_order_process)
        QShortcut(QKeySequence('R'), self, self.update_displays)
        QShortcut(QKeySequence('E'), self, self.edit_order_process)
        QShortcut(QKeySequence('Shift+P'), self, self.toggle_order_print)
        QShortcut(QKeySequence('P'), self, self.print_select_order)

    def update_shortcuts_widget(self):
        self.shortcuts_widget.setHtml("""
            <html>
            <body>
            <h2 style="margin-top: 0;">Key Shortcuts</h2>
            <p>N: New Order</p>
            <p>U: Update Order</p>
            <p>R: Refresh Display</p>
            <p>E: Edit Order</p>
            <p>Shift+P: Toggle Order Print</p>
            <p>P: Print Selected Order</p>
            </body>
            </html>
        """)

    def new_order_process(self):
        dialog = NewOrderDialog(self)
        if dialog.exec_():
            self.orders_list.append(dialog.new_order)
            self.update_displays()

    def update_order_process(self):
        dialog = UpdateOrderDialog(self)
        dialog.exec_()
        self.update_displays()

    def edit_order_process(self):
        dialog = EditOrderDialog(self)
        dialog.exec_()
        self.update_displays()

    def update_displays(self):

        def _update_manager_display():
            self.clear_layout(self.main_layout)

            header_text = "Order # | Time | Customer Name | Food Items | Drink Items | Card or Cash | Total Price"
            header_label = QLabel(header_text)
            header_label.setFont(QFont("Franklin Gothic Demi", 10, QFont.Bold))
            self.main_layout.addWidget(header_label)

            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)
            scroll_widget = QWidget()
            scroll_layout = QVBoxLayout(scroll_widget)

            all_orders = sorted(self.orders_list, key=lambda order: order.order_number)
            for order in all_orders:
                order_frame = QFrame()
                order_frame.setFrameStyle(QFrame.Box | QFrame.Plain)
                order_frame.setStyleSheet(f"background-color: {get_color_for_status(order.status)};")

                order_layout = QHBoxLayout(order_frame)

                total_cost_display = f"{order.total_cost:.2f}€" if isinstance(order.total_cost,
                                                                              (int, float)) else order.total_cost
                order_info = (
                    f"{order.order_number} | "
                    f"{order.creation_time.strftime('%H:%M:%S')} | "
                    f"{order.customer_name or 'N/A'} | "
                    f"{', '.join([f'{item} x {quantity}' for item, quantity in order.food_items.items()])} | "
                    f"{', '.join([f'{item} x {quantity}' for item, quantity in order.bar_items.items()])} | "
                    f"{order.card_or_cash} | "
                    f"{total_cost_display} | "
                )
                order_label = QLabel(order_info)
                order_layout.addWidget(order_label)

                scroll_layout.addWidget(order_frame)

            scroll_area.setWidget(scroll_widget)
            self.main_layout.addWidget(scroll_area)

            QTimer.singleShot(100, lambda: scroll_area.verticalScrollBar().setValue(
                scroll_area.verticalScrollBar().maximum()))

        def get_color_for_status(status):
            if status == "active":
                return "red"
            elif status == "completed":
                return "yellow"
            elif status == "picked up":
                return "green"
            elif status == "bar":
                return "blue"

        _update_manager_display()
        self.kitchen_window.update_display([order for order in self.orders_list if order.status == "active"])
        # self.send_update_to_rpi()
        self.save_orders_to_excel()

    def load_or_create_workbook(self):
        file_path = "orders.xlsx"
        if os.path.exists(file_path):
            self.workbook = load_workbook(file_path)
            self.worksheet = self.workbook.active
            print("Workbook loaded.")
            self.read_orders_from_excel()
        else:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Orders"
            headers = ["Order Number", "Creation Time", "Customer Name",
                       "Food Items", "Bar Items", "card_or_cash", "Total Cost", "Status"]
            self.worksheet.append(headers)
            self.workbook.save("orders.xlsx")
            print("New workbook created.")

    def read_orders_from_excel(self):
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            order_number, creation_time, customer_name, food_items, bar_items, c_or_c, total_cost, status = row

            creation_time = datetime.datetime.strptime(creation_time, '%Y-%m-%d %H:%M:%S')

            food_dict = {}
            bar_dict = {}
            if food_items:
                food_item_list = food_items.split(", ")
                for item in food_item_list:
                    name, quantity = item.split(" x")
                    food_dict[name] = int(quantity)

            if bar_items:
                bar_list = bar_items.split(", ")
                for item in bar_list:
                    name, quantity = item.split(" x")
                    bar_dict[name] = int(quantity)

            order = Order(
                order_number=order_number,
                food_items=food_dict,
                bar_items=bar_dict,
                customer_name=customer_name,
                creation_time=creation_time,
                card_or_cash=c_or_c,
                total_cost=total_cost,
                status=status
            )

            self.orders_list.append(order)

    def save_orders_to_excel(self):
        self.worksheet.delete_rows(2, self.worksheet.max_row)

        all_orders = self.orders_list
        all_orders.sort(key=lambda order: order.order_number)

        for order in all_orders:
            order_number = order.order_number
            creation_time = order.creation_time.strftime('%Y-%m-%d %H:%M:%S')
            customer_name = order.customer_name or "N/A"
            food_items = ", ".join([f"{item} x{quantity}" for item, quantity in order.food_items.items()])
            bar_items = ", ".join([f"{item} x{quantity}" for item, quantity in order.bar_items.items()])
            c_or_c = order.card_or_cash
            total_cost = order.total_cost if isinstance(order.total_cost, (int, float)) else str(order.total_cost)
            status = order.status

            self.worksheet.append([order_number, creation_time, customer_name,
                                   food_items, bar_items, c_or_c, total_cost, status])

        self.workbook.save("orders.xlsx")

    def print_select_order(self):
        dialog = PrintOrderDialog(self)
        dialog.exec_()

    def clear_layout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.setParent(None)
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())

    def toggle_order_print(self):
        NewOrderDialog.order_print = not NewOrderDialog.order_print
        self.update_info_widget()

    def update_info_widget(self):
        status = "Enabled" if NewOrderDialog.order_print else "Disabled"
        connection_status = "Connected" if self.rpi_connected else "Disconnected"
        connection_style = "" if self.rpi_connected else "background-color: red;"

        self.info_widget.setHtml(f"""
            <html>
            <body>
            <h2>Information Panel</h2>
            <p>Order Print: {status}</p>
            <p style='{connection_style}'>RPi: {connection_status}</p>
            </body>
            </html>
        """)

    def start_ping(self):
        self.ping_thread = PingThread(self.rpi_ip, self.rpi_port)
        self.ping_thread.ping_result.connect(self.handle_ping_result)
        self.ping_thread.start()

        # Set a timeout for the thread
        QTimer.singleShot(2000, self.ping_thread.quit)

    def handle_ping_result(self, connected):
        self.rpi_connected = connected
        self.update_info_widget()

    def send_update_to_rpi(self):
        active_order_numbers = [order for order in self.orders_list if order.status == "active"]
        completed_order_numbers = [order for order in self.orders_list if order.status == "completed"]

        data = {
            'active_orders': active_order_numbers,
            'completed_orders': completed_order_numbers
        }

        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.connect((self.rpi_ip, self.rpi_port))
                s.sendall(json.dumps(data).encode())
            self.rpi_connected = True
        except Exception as e:
            print(f"Failed to send update to Raspberry Pi: {e}")
            self.rpi_connected = False
        self.update_info_widget()


class PingThread(QThread):
    ping_result = pyqtSignal(bool)

    def __init__(self, rpi_ip, rpi_port):
        super().__init__()
        self.rpi_ip = rpi_ip
        self.rpi_port = rpi_port

    def run(self):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.setblocking(False)
                try:
                    s.connect((self.rpi_ip, self.rpi_port))
                except BlockingIOError:
                    pass

                readable, writable, _ = select.select([], [s], [], 1.0)

                if writable:
                    s.sendall(b'ping')
                    readable, _, _ = select.select([s], [], [], 1.0)
                    if readable:
                        response = s.recv(1024)
                        if response == b'pong':
                            self.ping_result.emit(True)
                            return

                self.ping_result.emit(False)
        except Exception as e:
            print(f"Error pinging RPi: {e}")
            self.ping_result.emit(False)


class KitchenWindow(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.setWindowTitle("Kitchen Display")
        self.setGeometry(400, 400, 1200, 800)  # Increased size to accommodate columns

        self.main_layout = QVBoxLayout(self)

    def update_display(self, active_orders):
        self.clear_layout(self.main_layout)
        active_orders.sort(key=lambda order: order.order_number)
        h_layout = QHBoxLayout()
        self.main_layout.addLayout(h_layout)

        left_column = QVBoxLayout()
        middle_column = QVBoxLayout()
        right_column = QVBoxLayout()
        h_layout.addLayout(left_column)
        h_layout.addLayout(middle_column)
        h_layout.addLayout(right_column)

        total_height = 0
        for i, order in enumerate(active_orders):
            order_widget, order_height = self.create_order_widget(order)
            total_height += order_height
            if total_height < 980:
                left_column.addWidget(order_widget)
            elif i < 1960:
                middle_column.addWidget(order_widget)
            elif i < 2940:
                right_column.addWidget(order_widget)
        left_column.addStretch(1)
        middle_column.addStretch(1)
        right_column.addStretch(1)

    def create_order_widget(self, order):
        order_widget = QFrame()
        order_widget.setStyleSheet("background-color: rgba(255, 0, 0, 255); padding: 2px;")
        order_layout = QVBoxLayout(order_widget)
        order_layout.setContentsMargins(5, 5, 5, 5)
        order_items = "\n ".join([f"{item} x{quantity}" for item, quantity in order.food_items.items()])
        order_info = (
            f"#{order.order_number}\n"
            f"{order_items}"
        )
        order_label = QLabel(order_info)
        order_label.setFont(QFont("", 30))
        order_label.setAlignment(Qt.AlignCenter)
        order_layout.addWidget(order_label)
        widget_height = 70 + 70 * len(order.food_items)
        order_widget.setFixedHeight(widget_height)
        return order_widget, widget_height

    def clear_layout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.setParent(None)
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())


class CustomerWindow(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.setWindowTitle("Customer View")
        self.setGeometry(300, 300, 1920, 1080)
        self.background = None
        self.scaled_background = None
        self.load_background()
        self.init_ui()

    def load_background(self):
        background_path = os.path.join(os.path.dirname(__file__), "background.jpg")
        if os.path.exists(background_path):
            self.background = QImage(background_path)
            if self.background.isNull():
                QMessageBox.warning(self, "Error", f"Failed to load background image: {background_path}")
            else:
                print(f"Background loaded. Original size: {self.background.size()}")
        else:
            QMessageBox.warning(self, "Error", f"Background image not found: {background_path}")

    def init_ui(self):
        # Main layout
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Create splitter
        self.splitter = QSplitter(Qt.Horizontal)
        self.splitter.setStyleSheet("QSplitter::handle { background-color: rgba(255, 255, 255, 100); }")
        main_layout.addWidget(self.splitter)

        # Create active orders widget
        self.active_widget = QWidget()
        self.active_widget.setAttribute(Qt.WA_TranslucentBackground)
        self.active_layout = QVBoxLayout(self.active_widget)
        self.active_layout.setAlignment(Qt.AlignTop)
        self.active_layout.setSpacing(10)
        self.splitter.addWidget(self.active_widget)

        # Create completed orders widget
        self.completed_widget = QWidget()
        self.completed_widget.setAttribute(Qt.WA_TranslucentBackground)
        self.completed_layout = QVBoxLayout(self.completed_widget)
        self.completed_layout.setAlignment(Qt.AlignTop)
        self.completed_layout.setSpacing(10)
        self.splitter.addWidget(self.completed_widget)

        # Set initial sizes
        self.splitter.setSizes([self.width() // 2, self.width() // 2])

        # Create header labels
        self.active_header = self.create_header_label("Küpseb")
        self.completed_header = self.create_header_label("Valmis")

        # Add headers to layouts
        self.active_layout.addWidget(self.active_header)
        self.completed_layout.addWidget(self.completed_header)

        # Create scroll areas for orders
        self.active_scroll = self.create_scroll_area()
        self.completed_scroll = self.create_scroll_area()

        self.active_layout.addWidget(self.active_scroll)
        self.completed_layout.addWidget(self.completed_scroll)

    def create_header_label(self, text):
        label = QLabel(text)
        label.setFont(QFont("Bebas neue", 150, QFont.Bold))
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("color: white; background-color: rgba(0, 0, 0, 100); border-radius: 10px; padding: 5px;")
        label.setFixedHeight(200)
        return label

    def paintEvent(self, event):
        painter = QPainter(self)
        if self.background and not self.background.isNull():
            if self.scaled_background is None or self.scaled_background.size() != self.size():
                self.scale_background()
            painter.drawImage(self.rect(), self.scaled_background)
        else:
            painter.fillRect(self.rect(), QColor(200, 200, 200))  # Light gray background

    def create_scroll_area(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        scroll_content = QWidget()
        scroll_content.setStyleSheet("background: transparent;")
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setAlignment(Qt.AlignTop)
        scroll_layout.setSpacing(5)
        scroll.setWidget(scroll_content)
        return scroll

    def scale_background(self):
        if self.background and not self.background.isNull():
            target_size = self.size()
            scaled_size = QSize(self.background.width(), self.background.height())
            scaled_size.scale(target_size, Qt.KeepAspectRatioByExpanding)

            self.scaled_background = self.background.scaled(
                scaled_size,
                Qt.KeepAspectRatioByExpanding,
                Qt.SmoothTransformation
            )

            if self.scaled_background.width() > target_size.width() or self.scaled_background.height() > target_size.height():
                x = (self.scaled_background.width() - target_size.width()) // 2
                y = (self.scaled_background.height() - target_size.height()) // 2
                self.scaled_background = self.scaled_background.copy(x, y, target_size.width(), target_size.height())

            print(f"Scaled background size: {self.scaled_background.size()}")

    def update_display(self, active_orders, completed_orders):
        self.update_orders(self.active_scroll, active_orders)
        self.update_orders(self.completed_scroll, completed_orders)

    def update_orders(self, scroll_area, orders):
        scroll_content = scroll_area.widget()
        layout = scroll_content.layout()

        self.clear_layout(layout)

        h_layout = QHBoxLayout()
        layout.addLayout(h_layout)

        left_column = QVBoxLayout()
        middle_column = QVBoxLayout()
        right_column = QVBoxLayout()
        h_layout.addLayout(left_column)
        h_layout.addLayout(middle_column)
        h_layout.addLayout(right_column)

        for i, order in enumerate(orders):
            order_widget = self.create_order_widget(order)
            if i < 5:
                left_column.addWidget(order_widget)
            elif i < 10:
                middle_column.addWidget(order_widget)
            elif i < 15:
                right_column.addWidget(order_widget)

        left_column.addStretch(1)
        middle_column.addStretch(1)
        right_column.addStretch(1)

    def create_order_widget(self, order):
        order_widget = QFrame()
        order_widget.setStyleSheet("background-color: rgba(255, 255, 255, 200); border-radius: 10px; padding: 2px;")
        order_layout = QVBoxLayout(order_widget)
        order_layout.setContentsMargins(10, 10, 10, 10)
        order_label = QLabel(f"{str(order.order_number).zfill(3)}")
        order_label.setFont(QFont("Bebas neue", 100))
        order_label.setAlignment(Qt.AlignCenter)
        order_layout.addWidget(order_label)
        order_widget.setFixedHeight(160)  # Set a fixed height for consistency
        return order_widget

    def clear_layout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.setParent(None)
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.scaled_background = None  # Force rescaling on next paint
        self.update()

    def showEvent(self, event):
        super().showEvent(event)
        self.scale_background()  # Ensure background is scaled when window is shown


class NewOrderDialog(QDialog):

    order_print = False

    def __init__(self, kitchenmanager=None):
        super().__init__(kitchenmanager)
        self.kitchenManager = kitchenmanager
        self.setWindowTitle("New Order")
        self.layout = QVBoxLayout(self)

        self.new_order = Order()

        self.init_ui()

    def init_ui(self):
        self.prompt_label = QLabel(Order.get_prompt_text())
        self.layout.addWidget(self.prompt_label)

        self.order_items_label = QLabel("Current Order:")
        self.layout.addWidget(self.order_items_label)

        self.payment_label = QLabel("Press Enter for normal payment, 'F' for free, 'C' for coupon")
        self.payment_label.hide()
        self.layout.addWidget(self.payment_label)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Enter customer name")
        self.name_input.hide()
        self.layout.addWidget(self.name_input)

        self.card_or_cash_label = QLabel("Enter: Card, C: cash")
        self.card_or_cash_label .hide()
        self.layout.addWidget(self.card_or_cash_label)

        self.set_shortcuts()

    def set_shortcuts(self):
        QShortcut(QKeySequence(Qt.Key_Return), self, lambda: self.ask_for_payment())
        for item, details in Order.combined_items_info.items():
            QShortcut(QKeySequence(details['key']), self, lambda ite=item: self.add_item(ite))
            QShortcut(QKeySequence(f"Shift+{details['key']}"), self, lambda ite=item: self.remove_item(ite))

    def add_item(self, item):
        Order.add_item(item, self.new_order)
        self.order_items_label.setText(Order.get_items_display(self.new_order))

    def remove_item(self, item):
        Order.remove_item(item, self.new_order)
        self.order_items_label.setText(Order.get_items_display(self.new_order))

    def ask_for_payment(self):
        self.remove_shortcuts([QKeySequence('F'), QKeySequence('C'), QKeySequence(Qt.Key_Return)])
        QShortcut(QKeySequence('F'), self, lambda: self.finalize_payment('free'))
        QShortcut(QKeySequence('C'), self, lambda: self.finalize_payment('coupon'))
        QShortcut(QKeySequence(Qt.Key_Return), self, lambda: self.finalize_payment('normal'))
        self.prompt_label.hide()
        self.order_items_label.hide()
        self.payment_label.show()
        self.setFocusPolicy(Qt.StrongFocus)
        self.setFocus()

    def finalize_payment(self, payment_type):
        if payment_type == 'free':
            self.new_order.total_cost = 'free'
            self.ask_for_name()
        elif payment_type == 'normal':
            self.new_order.total_cost = Order.calculate_total_price(self.new_order)
            self.cash_or_card()
        elif payment_type == 'coupon':
            self.new_order.total_cost = 'coupon'
            self.finish_order()


    def ask_for_name(self):
        self.remove_shortcuts([QKeySequence(Qt.Key_Return)])
        QShortcut(QKeySequence(Qt.Key_Return), self, lambda: self.finalize_ask_for_name())
        self.name_input.show()
        self.name_input.setFocus()
        self.payment_label.hide()
        self.prompt_label.hide()
        self.order_items_label.hide()

    def finalize_ask_for_name(self):
        self.new_order.customer_name = self.name_input.text()
        self.finish_order()

    def cash_or_card(self):
        self.remove_shortcuts([QKeySequence('C'), QKeySequence(Qt.Key_Return)])
        QShortcut(QKeySequence(Qt.Key_Return), self, lambda: self.finalize_cash_or_card('card'))
        QShortcut(QKeySequence('C'), self, lambda: self.finalize_cash_or_card('cash'))
        self.payment_label.hide()
        self.card_or_cash_label.show()

    def finalize_cash_or_card(self, c_or_c):
        self.new_order.card_or_cash = c_or_c
        self.finish_order()

    def finish_order(self):
        if self.order_print:
            print_order(self.new_order)
        if not self.new_order.food_items:
            self.new_order.status = "bar"
        self.accept()

    def remove_shortcuts(self, key_sequences):
        if not isinstance(key_sequences, list):
            key_sequences = [key_sequences]

        removed = False
        for key_sequence in key_sequences:
            for child in self.children():
                if isinstance(child, QShortcut) and child.key() == key_sequence:
                    child.setEnabled(False)
                    child.deleteLater()
                    removed = True
        return removed


class UpdateOrderDialog(QDialog):
    def __init__(self, kitchenmanager=None):
        super().__init__(kitchenmanager)
        self.kitchenManager = kitchenmanager
        self.setWindowTitle("Update Order")
        self.layout = QVBoxLayout(self)

        self.init_ui()

    def init_ui(self):
        self.order_input = QLineEdit()
        self.order_input.setPlaceholderText("Enter order number and press Enter")
        self.layout.addWidget(self.order_input)
        self.order_input.returnPressed.connect(self.process_order)

    def process_order(self):
        try:
            order_number = int(self.order_input.text())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid order number.")
            return

        for order in self.kitchenManager.orders_list:
            if order.order_number == order_number:
                if order.status == "active":
                    order.status = "completed"
                    self.accept()
                    break
                elif order.status == "completed":
                    order.status = "picked up"
                    self.accept()
                    break
                elif order.status == "picked up":
                    QMessageBox.information(self, "Order Already Done",
                                            f"Order {order_number} is already in the 'done' state.")
                    self.accept()
                    break
                elif order.status == "bar":
                    self.accept()
                    break
        else:
            QMessageBox.warning(self, "Order Not Found", f"No order found with number {order_number}")
            self.accept()


class EditOrderDialog(QDialog):
    def __init__(self, kitchen_manager=None):
        super().__init__(kitchen_manager)
        self.kitchenManager = kitchen_manager
        self.setWindowTitle("Edit Order")
        self.layout = QVBoxLayout(self)
        self.selected_order = None
        self.edit_mode = None

        self.init_ui()

    def init_ui(self):
        # Order number input
        self.edit_instructions = QLabel(Order.get_prompt_text())
        self.edit_instructions.hide()
        self.layout.addWidget(self.edit_instructions)
        self.order_input = QLineEdit()
        self.order_input.setPlaceholderText("Enter order number and press Enter")
        self.layout.addWidget(self.order_input)
        self.order_input.returnPressed.connect(self.find_order)

        # Instructions label
        self.instructions_label = QLabel("n: change name, p: change payment, s: change status, i: change food_items")
        self.instructions_label.hide()
        self.layout.addWidget(self.instructions_label)

        # Edit area
        self.edit_area = QTextEdit()
        self.edit_area.hide()
        self.layout.addWidget(self.edit_area)

        self.order_items_label = QLabel("Current Order:")
        self.order_items_label.hide()
        self.layout.addWidget(self.order_items_label)

        self.payment_instructions = QLabel("Press Enter for normal payment, 'F' for free, 'C' for coupon")
        self.payment_instructions.hide()
        self.layout.addWidget(self.payment_instructions)

        self.status_label = QLabel("a: active, c: completed, p: picked up")
        self.status_label.hide()
        self.layout.addWidget(self.status_label)

        # Set up shortcuts
        self.setup_shortcuts()

        # Set focus on order input
        self.order_input.setFocus()

    def setup_shortcuts(self):
        QShortcut(QKeySequence('N'), self, self.edit_name)
        QShortcut(QKeySequence('I'), self, self.edit_items)
        QShortcut(QKeySequence('P'), self, self.edit_payment)
        QShortcut(QKeySequence('S'), self, self.edit_status)
        QShortcut(QKeySequence('D'), self, self.delete_order)

    def find_order(self):
        try:
            order_number = int(self.order_input.text())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid order number.")
            return

        for order in self.kitchenManager.orders_list:
            if order.order_number == order_number:
                self.selected_order = order
                self.show_order_details()
                return

        QMessageBox.warning(self, "Order Not Found", f"No order found with number {order_number}")

    def show_order_details(self):
        self.order_input.hide()
        self.instructions_label.show()
        self.edit_area.show()
        self.edit_area.setPlainText(f"Order #{self.selected_order.order_number}\n"
                                    f"Name: {self.selected_order.customer_name}\n"
                                    f"Items: {', '.join([f'{item} x{qty}' for item, qty in self.selected_order.combined_order_items.items()])}\n"
                                    f"Payment: {self.selected_order.total_cost}")
        self.edit_area.setReadOnly(True)
        self.setFocus()

    def edit_name(self):
        new_name, ok = QInputDialog.getText(self, "Edit Name", "Enter new name:")
        if ok and new_name:
            self.selected_order.customer_name = new_name
            self.accept()

    def edit_payment(self):

        def set_shortcuts():
            QShortcut(QKeySequence(Qt.Key_Return), self, lambda: finalize_payment('normal'))
            QShortcut(QKeySequence('F'), self, lambda: finalize_payment('free'))
            QShortcut(QKeySequence('C'), self, lambda: finalize_payment('coupon'))

        def finalize_payment(payment_type):
            if payment_type == 'normal':
                self.selected_order.total_cost = Order.calculate_total_price(self.selected_order.food_items)
            elif payment_type == 'free':
                self.selected_order.total_cost = 'free'
            elif payment_type == 'coupon':
                self.selected_order.total_cost = 'coupon'

            self.accept()

        if not self.selected_order:
            return
        self.edit_area.hide()
        self.instructions_label.hide()
        self.payment_instructions.show()
        set_shortcuts()

    def edit_items(self):

        def set_shortcuts():
            for shortcut in self.findChildren(QShortcut):
                shortcut.setEnabled(False)
            QShortcut(QKeySequence(Qt.Key_Return), self, finalize_items)
            for item, details in Order.combined_items_info.items():
                QShortcut(QKeySequence(details['key']), self, lambda ite=item: add_item(ite))
                QShortcut(QKeySequence(f"Shift+{details['key']}"), self, lambda ite=item: remove_item(ite))

        def add_item(item):
            Order.add_item(item, self.selected_order)
            self.order_items_label.setText(Order.get_items_display(self.selected_order))

        def remove_item(item):
            Order.remove_item(item, self.selected_order)
            self.order_items_label.setText(Order.get_items_display(self.selected_order))

        def finalize_items():
            if isinstance(self.selected_order.total_cost, (int, float)):
                self.selected_order.total_cost = Order.calculate_total_price(self.selected_order)
            self.accept()

        self.order_items_label.setText(Order.get_items_display(self.selected_order))
        self.edit_instructions.show()
        self.order_items_label.show()
        self.instructions_label.hide()
        self.edit_area.hide()
        set_shortcuts()

    def edit_status(self):

        def change_status(status_type):
            self.selected_order.status = status_type
            self.accept()

        def set_shortcuts():
            for shortcut in self.findChildren(QShortcut):
                shortcut.setEnabled(False)

            QShortcut(QKeySequence('A'), self, lambda: change_status('active'))
            QShortcut(QKeySequence("C"), self, lambda: change_status('completed'))
            QShortcut(QKeySequence("P"), self, lambda: change_status('picked up'))

        self.edit_area.hide()
        self.instructions_label.hide()
        self.status_label.show()
        set_shortcuts()

    def delete_order(self):
        self.kitchenManager.orders_list.remove(self.selected_order)
        self.accept()


class PrintOrderDialog(QDialog):
    def __init__(self, kitchen_manager=None):
        super().__init__(kitchen_manager)
        self.KitchenManager = kitchen_manager
        self.setWindowTitle("Update Order")
        self.layout = QVBoxLayout(self)

        self.init_ui()

    def init_ui(self):
        self.order_input = QLineEdit()
        self.order_input.setPlaceholderText("Enter order number and press Enter")
        self.layout.addWidget(self.order_input)
        self.order_input.returnPressed.connect(self.process_order)

    def process_order(self):
        try:
            order_number = int(self.order_input.text())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid order number.")
            return

        for order in self.KitchenManager.order_list:
            if order.order_number == order_number:
                print_order(order)
                self.accept()
                break
        else:
            QMessageBox.warning(self, "Order Not Found", f"No order found with number {order_number}")

        self.order_input.clear()


def print_order(order):
    if order.bar_items:
        print(order.bar_items)
        print_text('Bari joogid:', size=2)
        for item, count in order.bar_items.items():
            print(f"{item} x{count}")
            print_text(f"{item} x{count}")
            print_text('\n\n\n\n\n\n\n\n')

    if order.food_items:
        print_text('Sinu tellimus:', size=2)
        print('Sinu tellimus:')
        for item, count in order.food_items.items():
            print(f"{item} x{count}")
            print_text(f"{item} x{count}")
        print_text('Tellimuse nr:', size=2)
        print('Tellimuse nr:')
        print_image(f"{str(order.order_number).zfill(3)}")
        print(f"{str(order.order_number).zfill(3)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = KitchenManagerApp()
    window.show()
    sys.exit(app.exec_())